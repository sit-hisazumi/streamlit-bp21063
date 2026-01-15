import json
import os
import re
from datetime import datetime
from io import BytesIO

import openpyxl
import pandas as pd
import streamlit as st
from fpdf import FPDF

# ファイルパス
JSON_FILE = "data.json"
IMAGES_DIR = "images"
TEMPLATE_FILE = "templates/inspection_template.xlsx"


def ensure_directories():
    """必要なディレクトリを作成する"""
    if not os.path.exists(IMAGES_DIR):
        os.makedirs(IMAGES_DIR)


def load_parts_data():
    """JSONファイルから部品データを読み込む"""
    if not os.path.exists(JSON_FILE):
        return []

    with open(JSON_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data.get("parts", [])


def save_parts_data(parts):
    """部品データをJSONファイルに保存する"""
    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump({"parts": parts}, f, ensure_ascii=False, indent=2)


def save_part(part_data, image_file=None):
    """新しい部品を追加する（画像があれば保存）"""
    parts = load_parts_data()

    # 画像を保存
    if image_file is not None:
        ext = os.path.splitext(image_file.name)[1]
        image_filename = f"{part_data['id']}{ext}"
        image_path = os.path.join(IMAGES_DIR, image_filename)

        with open(image_path, "wb") as f:
            f.write(image_file.getbuffer())

        part_data["image_file"] = image_filename
    else:
        part_data["image_file"] = None

    parts.append(part_data)
    save_parts_data(parts)


def update_part(part_id, updated_data, image_file=None):
    """既存の部品を更新する（画像があれば保存）"""
    parts = load_parts_data()

    # 部品を検索
    part_index = None
    for idx, part in enumerate(parts):
        if part["id"] == part_id:
            part_index = idx
            break

    if part_index is None:
        return False

    # 画像を保存
    if image_file is not None:
        ext = os.path.splitext(image_file.name)[1]
        image_filename = f"{updated_data['id']}{ext}"
        image_path = os.path.join(IMAGES_DIR, image_filename)

        with open(image_path, "wb") as f:
            f.write(image_file.getbuffer())

        updated_data["image_file"] = image_filename
    else:
        # 画像ファイルが指定されていない場合は既存の画像を保持
        if "image_file" not in updated_data:
            updated_data["image_file"] = parts[part_index].get("image_file")

    # 部品を更新
    parts[part_index] = updated_data
    save_parts_data(parts)
    return True


def get_image_path(part):
    """部品の画像パスを取得する（存在する場合）"""
    if part.get("image_file"):
        path = os.path.join(IMAGES_DIR, part["image_file"])
        if os.path.exists(path):
            return path
    return None


def load_inspection_template():
    """Excelテンプレートから検査項目を読み込む"""
    if not os.path.exists(TEMPLATE_FILE):
        # デフォルトの検査項目
        return [
            {"no": 1, "item": "外観検査", "criteria": "傷・変形・錆なきこと"},
            {"no": 2, "item": "寸法検査（長さ）", "criteria": "100±0.5mm"},
            {"no": 3, "item": "寸法検査（幅）", "criteria": "50±0.3mm"},
            {"no": 4, "item": "寸法検査（厚さ）", "criteria": "10±0.1mm"},
            {"no": 5, "item": "硬度検査", "criteria": "HRC 58-62"},
            {"no": 6, "item": "動作確認", "criteria": "スムーズに動作すること"},
        ]

    wb = openpyxl.load_workbook(TEMPLATE_FILE)
    ws = wb.active

    items = []
    for row in range(6, 12):  # 6行目から11行目まで（6項目）
        no = ws.cell(row=row, column=1).value
        item = ws.cell(row=row, column=2).value
        criteria = ws.cell(row=row, column=3).value
        if no and item:
            items.append({"no": no, "item": item, "criteria": criteria or ""})

    return items


def extract_product_from_drawing_number(drawing_number):
    """
    図番から製品IDを抽出
    例: 【R】TUA60-BBBB-CCCC → TUA60
    """
    # 【R】を削除
    clean_number = drawing_number.replace("【R】", "").strip()

    # ハイフンで分割して最初の部分を取得
    if "-" in clean_number:
        return clean_number.split("-")[0]

    return clean_number


def parse_csv_file(uploaded_file):
    """
    CSVファイルをパースして部品データのリストを返す

    CSVフォーマット:
    - 1行目: ヘッダー
    - 2列目のみ値がある行: 製品カテゴリ (その後の行はこの製品用の部品)
    - 2,3,4列目に値がある行: 部品データ (品目, 図番, 品名)
    """
    parts_list = []
    current_product_name = None

    # CSVを読み込み
    df = pd.read_csv(uploaded_file, header=0, encoding='utf-8-sig')

    for idx, row in df.iterrows():
        # 列のインデックスで取得（0始まり）
        item_type = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
        drawing_number = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ""
        part_name = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ""

        # 品目のみの行 = 製品カテゴリ
        # 図番と品名が空（nanまたは空文字）の場合
        if item_type and (not drawing_number or drawing_number == 'nan') and (not part_name or part_name == 'nan'):
            current_product_name = item_type
            continue

        # 品目+図番+品名がある行 = 部品データ
        if item_type and drawing_number and part_name and drawing_number != 'nan' and part_name != 'nan':
            # 図番から製品IDを抽出
            product_id = extract_product_from_drawing_number(drawing_number)

            # 【R】を削除した図番をIDとして使用
            clean_id = drawing_number.replace("【R】", "").strip()

            part_data = {
                "id": clean_id,
                "name": part_name,
                "category": "未設定",
                "item_type": item_type,
                "inspection_items": ["未設定"],
                "cautions": ["未設定"],
                "storage": "未設定",
                "image_description": "検査箇所",
                "image_file": None,
                "required_products": []
            }

            # 製品情報を追加
            if product_id and current_product_name:
                part_data["required_products"].append({
                    "product_id": product_id,
                    "product_name": current_product_name,
                    "notes": ""
                })

            parts_list.append(part_data)

    return parts_list


def check_duplicates(parts_to_import, existing_parts):
    """
    インポート対象の部品と既存部品で重複をチェック
    """
    existing_ids = {part["id"] for part in existing_parts}
    duplicates = []
    unique_parts = []

    for part in parts_to_import:
        if part["id"] in existing_ids:
            duplicates.append(part)
        else:
            unique_parts.append(part)

    return unique_parts, duplicates


def import_parts_from_csv(parts_to_import, existing_parts, overwrite_duplicates=False):
    """
    CSVから読み込んだ部品をインポート
    """
    unique_parts, duplicates = check_duplicates(parts_to_import, existing_parts)

    success_count = 0
    skip_count = 0
    error_count = 0

    if overwrite_duplicates:
        # 重複する部品を上書き
        existing_dict = {part["id"]: part for part in existing_parts}
        for dup_part in duplicates:
            existing_dict[dup_part["id"]] = dup_part
            success_count += 1

        # ユニークな部品を追加
        for part in unique_parts:
            existing_dict[part["id"]] = part
            success_count += 1

        # 結果をリストに変換
        result_parts = list(existing_dict.values())
    else:
        # 重複をスキップ
        skip_count = len(duplicates)
        success_count = len(unique_parts)
        result_parts = existing_parts + unique_parts

    return result_parts, success_count, skip_count, error_count, duplicates


def auto_judge(item_no, result, criteria):
    """測定値から自動判定を行う"""
    if not result:
        return ""

    result = result.strip()

    # 項目1, 6は「OK」で合格
    if item_no in [1, 6]:
        if result.upper() == "OK":
            return "合格"
        elif result.upper() == "NG":
            return "不合格"
        return ""

    # 項目2-5は数値判定（範囲チェック）
    # 判定基準のパターン: "100±0.5mm", "HRC 58-62"
    try:
        # 測定値を数値に変換
        result_value = float(result.replace(",", "."))

        # ±形式の判定基準をパース（例: "100±0.5mm"）
        if "±" in criteria:
            import re
            match = re.search(r"([\d.]+)±([\d.]+)", criteria)
            if match:
                base = float(match.group(1))
                tolerance = float(match.group(2))
                if base - tolerance <= result_value <= base + tolerance:
                    return "合格"
                else:
                    return "不合格"

        # 範囲形式の判定基準をパース（例: "HRC 58-62"）
        if "-" in criteria:
            import re
            match = re.search(r"([\d.]+)-([\d.]+)", criteria)
            if match:
                min_val = float(match.group(1))
                max_val = float(match.group(2))
                if min_val <= result_value <= max_val:
                    return "合格"
                else:
                    return "不合格"

    except (ValueError, AttributeError):
        pass

    return ""


class JapanesePDF(FPDF):
    """日本語対応PDF"""

    def __init__(self):
        super().__init__()
        # 日本語フォントを追加
        font_path = "fonts/NotoSansJP-Regular.ttf"
        if os.path.exists(font_path):
            self.add_font("NotoSansJP", "", font_path)
            self.font_name = "NotoSansJP"
        else:
            self.font_name = "Helvetica"

    def header(self):
        self.set_font(self.font_name, "", 16)
        if self.font_name == "NotoSansJP":
            self.cell(0, 10, "部品検査表", align="C", new_x="LMARGIN", new_y="NEXT")
        else:
            self.cell(
                0, 10, "Inspection Report", align="C", new_x="LMARGIN", new_y="NEXT"
            )
        self.ln(5)


def generate_pdf(inspection_data, part_data):
    """検査結果をPDFに出力する"""
    pdf = JapanesePDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # 基本情報
    pdf.set_font(pdf.font_name, "", 10)

    pdf.cell(30, 8, "検査日:", border=1)
    pdf.cell(50, 8, inspection_data.get("date", ""), border=1)
    pdf.cell(30, 8, "検査者:", border=1)
    pdf.cell(50, 8, inspection_data.get("inspector", ""), border=1)
    pdf.ln()

    pdf.cell(30, 8, "部品ID:", border=1)
    pdf.cell(50, 8, part_data.get("id", ""), border=1)
    pdf.cell(30, 8, "部品名:", border=1)
    pdf.cell(50, 8, part_data.get("name", ""), border=1)
    pdf.ln()
    pdf.ln(5)

    # 検査項目テーブルヘッダー
    pdf.set_fill_color(68, 114, 196)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(10, 8, "No.", border=1, fill=True, align="C")
    pdf.cell(40, 8, "検査項目", border=1, fill=True, align="C")
    pdf.cell(45, 8, "判定基準", border=1, fill=True, align="C")
    pdf.cell(35, 8, "測定値/結果", border=1, fill=True, align="C")
    pdf.cell(20, 8, "判定", border=1, fill=True, align="C")
    pdf.cell(40, 8, "備考", border=1, fill=True, align="C")
    pdf.ln()

    # 検査項目データ
    pdf.set_text_color(0, 0, 0)
    for item in inspection_data.get("items", []):
        pdf.cell(10, 8, str(item.get("no", "")), border=1, align="C")
        pdf.cell(40, 8, item.get("item", "")[:15], border=1)
        pdf.cell(45, 8, item.get("criteria", "")[:18], border=1)
        pdf.cell(35, 8, item.get("result", ""), border=1, align="C")
        pdf.cell(20, 8, item.get("judgment", ""), border=1, align="C")
        pdf.cell(40, 8, item.get("note", "")[:15], border=1)
        pdf.ln()

    # 総合判定
    pdf.ln(5)
    pdf.set_fill_color(217, 226, 243)
    pdf.cell(40, 10, "総合判定:", border=1, fill=True, align="C")
    overall = inspection_data.get("overall_judgment", "")
    if overall == "合格":
        pdf.set_text_color(0, 128, 0)
    elif overall == "不合格":
        pdf.set_text_color(255, 0, 0)
    pdf.cell(60, 10, overall, border=1, align="C")
    pdf.set_text_color(0, 0, 0)

    return bytes(pdf.output())


# ディレクトリ作成
ensure_directories()

# ページ設定
st.set_page_config(
    page_title="部品検査箇所表示",
    page_icon="🔍",
    layout="wide"
)

# データ読み込み
parts_data = load_parts_data()

# カテゴリ一覧を取得
if parts_data:
    categories = ["すべて"] + sorted(
        list(set(part["category"] for part in parts_data))
    )
else:
    categories = ["すべて"]

# 製品一覧を取得（required_productsから抽出）
products_set = set()
for part in parts_data:
    for product in part.get("required_products", []):
        products_set.add(
            (product["product_id"], product["product_name"])
        )
products = ["すべて"] + sorted(
    [f"{pid} - {pname}" for pid, pname in products_set]
)

# セッション状態の初期化
if "selected_part" not in st.session_state:
    st.session_state.selected_part = None
if "show_add_form" not in st.session_state:
    st.session_state.show_add_form = False
if "show_inspection_form" not in st.session_state:
    st.session_state.show_inspection_form = False
if "inspection_results" not in st.session_state:
    st.session_state.inspection_results = {}

# Query parameter handling for navigation
query_params = st.query_params
current_view = query_params.get("view", "main")
selected_part_id_from_url = query_params.get("part_id", None)
selected_product_id_from_url = query_params.get("product_id", None)
preselected_part_id_for_inspection = query_params.get("selected_part_id", None)

# サイドバー（検索・フィルタ）
st.sidebar.title("🔍 検索・フィルタ")

# Home button if not on main view
if current_view != "main":
    if st.sidebar.button("🏠 ホームに戻る", width="stretch"):
        # Keep filter parameters but clear view parameters
        filters_to_keep = {}
        if "selected_product" in st.query_params:
            filters_to_keep["selected_product"] = st.query_params["selected_product"]
        if "search_query" in st.query_params:
            filters_to_keep["search_query"] = st.query_params["search_query"]
        if "selected_category" in st.query_params:
            filters_to_keep["selected_category"] = st.query_params["selected_category"]

        st.query_params.clear()
        for key, value in filters_to_keep.items():
            st.query_params[key] = value
        st.rerun()
    st.sidebar.markdown("---")

# 製品で絞り込み（URLパラメータから復元）
default_product_index = 0
if "selected_product" in st.query_params:
    saved_product = st.query_params["selected_product"]
    if saved_product in products:
        default_product_index = products.index(saved_product)

selected_product = st.sidebar.selectbox(
    "製品で絞り込み",
    products,
    index=default_product_index,
    help="特定の製品に必要な部品のみを表示"
)

# 製品選択が変わったらURLパラメータを更新
if selected_product != products[default_product_index] or "selected_product" not in st.query_params:
    st.query_params["selected_product"] = selected_product

# 検索クエリ（URLパラメータから復元）
default_search = st.query_params.get("search_query", "")
search_query = st.sidebar.text_input(
    "部品名・IDで検索",
    placeholder="例: ボルト, BLT-001",
    value=default_search
)

# 検索クエリが変わったらURLパラメータを更新
if search_query != default_search:
    if search_query:
        st.query_params["search_query"] = search_query
    elif "search_query" in st.query_params:
        del st.query_params["search_query"]

# カテゴリで絞り込み（URLパラメータから復元）
default_category_index = 0
if "selected_category" in st.query_params:
    saved_category = st.query_params["selected_category"]
    if saved_category in categories:
        default_category_index = categories.index(saved_category)

selected_category = st.sidebar.selectbox(
    "カテゴリで絞り込み",
    categories,
    index=default_category_index
)

# カテゴリ選択が変わったらURLパラメータを更新
if selected_category != categories[default_category_index] or "selected_category" not in st.query_params:
    st.query_params["selected_category"] = selected_category

# フィルタリング処理
filtered_parts = parts_data.copy()

# 製品による絞り込み
if selected_product != "すべて":
    product_id = selected_product.split(" - ")[0]
    filtered_parts = [
        part for part in filtered_parts
        if any(
            p["product_id"] == product_id
            for p in part.get("required_products", [])
        )
    ]

if search_query:
    filtered_parts = [
        part for part in filtered_parts
        if search_query.lower() in part["name"].lower()
        or search_query.lower() in part["id"].lower()
    ]

if selected_category != "すべて":
    filtered_parts = [
        part for part in filtered_parts
        if part["category"] == selected_category
    ]

# サイドバーに検索結果数を表示
st.sidebar.markdown("---")
st.sidebar.info(f"該当部品: {len(filtered_parts)} 件")

# 部品追加ボタン（ページ遷移に変更）
st.sidebar.markdown("---")
if st.sidebar.button("➕ 新規部品を追加", width="stretch"):
    st.query_params["view"] = "add_part"
    # Keep current filters
    if selected_product != "すべて":
        st.query_params["selected_product"] = selected_product
    if search_query:
        st.query_params["search_query"] = search_query
    if selected_category != "すべて":
        st.query_params["selected_category"] = selected_category
    st.rerun()

# 検査表ボタン（ページ遷移に変更）
if st.sidebar.button("📋 検査表を作成", width="stretch"):
    st.query_params["view"] = "inspection_form"
    # Keep current filters
    if selected_product != "すべて":
        st.query_params["selected_product"] = selected_product
    if search_query:
        st.query_params["search_query"] = search_query
    if selected_category != "すべて":
        st.query_params["selected_category"] = selected_category
    st.rerun()


# ============================================================
# View Functions
# ============================================================

def show_part_details_page(part_id, parts_data):
    """Display detailed part information page"""
    # Find the selected part
    part_data = next((p for p in parts_data if p["id"] == part_id), None)

    if not part_data:
        st.error(f"部品ID '{part_id}' が見つかりません。")
        if st.button("ホームに戻る"):
            st.query_params.clear()
            st.rerun()
        return

    # Title with buttons
    title_col, btn_col1, btn_col2 = st.columns([3, 1, 1])
    with title_col:
        st.title(f"📋 {part_data['name']}")
    with btn_col1:
        st.markdown("<br>", unsafe_allow_html=True)  # Add spacing
        if st.button("📋 検査表作成", key="create_inspection_btn", use_container_width=True):
            st.query_params["view"] = "inspection_form"
            st.query_params["selected_part_id"] = part_id
            st.rerun()
    with btn_col2:
        st.markdown("<br>", unsafe_allow_html=True)  # Add spacing
        if st.button("✏️ 編集", key="edit_part_btn", type="primary", use_container_width=True):
            st.query_params["view"] = "edit_part"
            st.query_params["part_id"] = part_id
            st.rerun()
    st.markdown("---")

    # 2列レイアウトで詳細表示
    col1, col2 = st.columns([1, 1])

    with col1:
        st.markdown(f"### {part_data['name']}")
        st.markdown(f"**部品番号:** {part_data['id']}")
        st.markdown(f"**カテゴリ:** {part_data['category']}")
        st.markdown(f"**保管場所:** 📍 {part_data['storage']}")

        # 必須製品
        if part_data.get("required_products"):
            st.markdown("#### 🏭 このパーツを使用する製品")
            for product in part_data["required_products"]:
                st.info(
                    f"**{product['product_id']}** - "
                    f"{product['product_name']}"
                )
                if product.get('notes'):
                    st.caption(f"用途: {product['notes']}")

        # 検査項目
        st.markdown("#### ✅ 検査項目")
        for item in part_data["inspection_items"]:
            st.markdown(f"- {item}")

        # 注意点
        st.markdown("#### ⚠️ 注意点")
        for caution in part_data["cautions"]:
            st.warning(caution)

    with col2:
        # 検査箇所画像
        st.markdown("#### 🖼️ 検査箇所イメージ")

        image_path = get_image_path(part_data)
        if image_path:
            # 画像がある場合は表示
            st.image(
                image_path,
                caption=part_data.get("image_description", "検査箇所"),
                width="stretch"
            )
        else:
            # プレースホルダー表示
            st.markdown(
                f"""
                <div style="
                    background-color: #f5f5f5;
                    border: 2px dashed #ccc;
                    border-radius: 10px;
                    padding: 60px 20px;
                    text-align: center;
                    color: #666;
                ">
                    <div style="font-size: 48px;">🔍</div>
                    <div style="margin-top: 10px; font-weight: bold;">
                        {part_data.get('image_description', '検査箇所')}
                    </div>
                    <div style="margin-top: 5px; font-size: 12px; color: #999;">
                        ※ 画像が登録されていません
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )


def show_product_details_page(product_id, product_name, parts_data):
    """Display detailed product information page"""
    # Find all parts that use this product
    related_parts = [
        part for part in parts_data
        if any(
            p["product_id"] == product_id
            for p in part.get("required_products", [])
        )
    ]

    if not related_parts:
        st.error(f"製品 '{product_name}' に関連する部品が見つかりません。")
        if st.button("ホームに戻る"):
            st.query_params.clear()
            st.rerun()
        return

    # Title
    st.title(f"🏭 {product_name}")
    st.markdown(f"**製品ID:** {product_id}")
    st.info(f"この製品には **{len(related_parts)}個** の部品が必要です。")
    st.markdown("---")

    # Display all related parts
    st.subheader("📦 必要な部品一覧")

    for part in related_parts:
        # Find the product note for this specific part
        product_note = next(
            (
                p["notes"] for p in part.get("required_products", [])
                if p["product_id"] == product_id
            ),
            ""
        )

        with st.expander(
            f"**{part['id']}** - {part['name']} "
            f"({part['category']})",
            expanded=False
        ):
            col1, col2 = st.columns([1, 1])

            with col1:
                if product_note:
                    st.markdown(f"**用途:** {product_note}")
                st.markdown(f"**保管場所:** 📍 {part['storage']}")

                # 検査項目
                st.markdown("**✅ 検査項目:**")
                for item in part["inspection_items"]:
                    st.markdown(f"- {item}")

                # 注意点
                st.markdown("**⚠️ 注意点:**")
                for caution in part["cautions"]:
                    st.caption(f"• {caution}")

            with col2:
                image_path = get_image_path(part)
                if image_path:
                    st.image(
                        image_path,
                        caption=part.get("image_description", "検査箇所"),
                        width="stretch"
                    )
                else:
                    st.caption(part.get("image_description", "検査箇所"))

            # Button to view full part details
            if st.button(
                "詳細を見る",
                key=f"view_part_{part['id']}",
                width="stretch"
            ):
                # Keep the product filter when navigating to part details
                st.query_params["view"] = "part_details"
                st.query_params["part_id"] = part["id"]
                if "selected_product" in st.query_params:
                    # Keep the current product filter
                    pass  # Already in query params
                st.rerun()


def show_add_part_page(parts_data):
    """Display add part page"""
    st.title("➕ 新規部品登録")
    st.markdown("---")

    # タブで手動登録とCSV一括登録を切り替え
    tab1, tab2 = st.tabs(["✍️ 手動登録", "📁 CSV一括登録"])

    with tab1:
        # 手動登録フォーム
        with st.form("add_part_form"):
            col1, col2 = st.columns(2)

            with col1:
                new_id = st.text_input("部品ID *", placeholder="例: BLT-002")
                new_name = st.text_input("部品名 *", placeholder="例: 六角ボルト M12")
                new_category = st.text_input("カテゴリ *", placeholder="例: 締結部品")
                new_storage = st.text_input(
                    "保管場所 *", placeholder="例: A棟-1F-棚番号A-15"
                )

            with col2:
                new_inspection = st.text_area(
                    "検査項目 *（1行に1項目）",
                    placeholder="ねじ山の損傷確認\n頭部の変形確認\n表面の錆確認",
                    height=100
                )
                new_cautions = st.text_area(
                    "注意点（1行に1項目）",
                    placeholder="トルク管理が重要\n再使用回数に注意",
                    height=100
                )
                new_image_desc = st.text_input(
                    "検査箇所イメージの説明",
                    placeholder="例: ボルト頭部・ねじ山部の検査ポイント"
                )
                new_required_products = st.text_area(
                    "必須製品（任意、1行に1製品）",
                    placeholder="TUA60|TUA60 アセンブリ|主軸固定用\nTUA70|TUA70 ユニット|予備用",
                    height=80,
                    help="形式: 製品ID|製品名|用途（パイプ区切り）"
                )

            # 画像アップロード
            uploaded_image = st.file_uploader(
                "検査箇所の画像（任意）",
                type=["png", "jpg", "jpeg"],
                help="PNG, JPG, JPEG形式の画像をアップロードできます"
            )

            submitted = st.form_submit_button("登録", width="stretch")

            if submitted:
                # バリデーション
                if not new_id or not new_name or not new_category or not new_storage:
                    st.error("必須項目（*）を入力してください。")
                elif any(part["id"] == new_id for part in parts_data):
                    st.error(f"部品ID '{new_id}' は既に存在します。")
                elif not new_inspection.strip():
                    st.error("検査項目を1つ以上入力してください。")
                else:
                    # 必須製品のパース
                    required_products = []
                    if new_required_products.strip():
                        for line in new_required_products.split("\n"):
                            if line.strip():
                                parts_info = [p.strip() for p in line.split("|")]
                                if len(parts_info) >= 2:
                                    product = {
                                        "product_id": parts_info[0],
                                        "product_name": parts_info[1],
                                        "notes": (
                                            parts_info[2]
                                            if len(parts_info) >= 3
                                            else ""
                                        )
                                    }
                                    required_products.append(product)

                    # 新規部品データを作成
                    new_part = {
                        "id": new_id,
                        "name": new_name,
                        "category": new_category,
                        "inspection_items": [
                            item.strip() for item in new_inspection.split("\n")
                            if item.strip()
                        ],
                        "cautions": [
                            item.strip() for item in new_cautions.split("\n")
                            if item.strip()
                        ] if new_cautions.strip() else ["特になし"],
                        "storage": new_storage,
                        "image_description": (
                            new_image_desc if new_image_desc else "検査箇所"
                        ),
                        "required_products": required_products
                    }

                    # JSONに保存（画像も含む）
                    save_part(new_part, uploaded_image)
                    st.success(f"部品 '{new_name}' を登録しました！")

                    # ホームに戻るボタンを表示
                    if st.button("🏠 ホームに戻る", type="primary"):
                        st.query_params.clear()
                        st.rerun()

    with tab2:
        # CSV一括登録フォーム
        st.markdown("#### 📁 CSVファイルから部品を一括登録")
        st.caption(
            "CSVフォーマット: 2列目=品目、3列目=図番、4列目=品名。"
            "品目のみの行は製品カテゴリを表します。"
        )

        # セッション状態の初期化
        if "csv_parsed_parts" not in st.session_state:
            st.session_state.csv_parsed_parts = []
        if "csv_import_result" not in st.session_state:
            st.session_state.csv_import_result = None

        # CSVファイルアップロード
        uploaded_csv = st.file_uploader(
            "CSVファイルを選択",
            type=["csv"],
            help="部品情報が記載されたCSVファイルをアップロードしてください",
            key="csv_uploader"
        )

        if uploaded_csv is not None:
            try:
                # CSVをパース
                parsed_parts = parse_csv_file(uploaded_csv)
                st.session_state.csv_parsed_parts = parsed_parts

                if len(parsed_parts) > 0:
                    st.success(f"✅ {len(parsed_parts)} 件の部品データを読み込みました")

                    # プレビューテーブル
                    st.markdown("#### 📋 プレビュー")
                    preview_data = []
                    for part in parsed_parts[:10]:  # 最初の10件を表示
                        preview_data.append({
                            "部品ID": part["id"],
                            "部品名": part["name"],
                            "品目": part.get("item_type", ""),
                            "製品": (
                                part["required_products"][0]["product_name"]
                                if part["required_products"]
                                else ""
                            )
                        })

                    st.dataframe(preview_data, use_container_width=True)

                    if len(parsed_parts) > 10:
                        st.caption(f"...他 {len(parsed_parts) - 10} 件")

                    # 重複チェック
                    unique_parts, duplicates = check_duplicates(
                        parsed_parts, parts_data
                    )

                    if duplicates:
                        st.warning(
                            f"⚠️ {len(duplicates)} 件の重複する部品IDがあります"
                        )
                        with st.expander("重複する部品ID一覧"):
                            for dup in duplicates:
                                st.markdown(f"- {dup['id']}: {dup['name']}")

                    # インポート設定
                    st.markdown("#### ⚙️ インポート設定")
                    overwrite = st.checkbox(
                        "重複する部品を上書きする",
                        value=False,
                        help="チェックすると、既存の部品データを上書きします"
                    )

                    # インポートボタン
                    if st.button(
                        f"📥 {len(parsed_parts)} 件の部品をインポート",
                        type="primary",
                        width="stretch"
                    ):
                        # インポート実行
                        result_parts, success, skip, error, dup_list = (
                            import_parts_from_csv(
                                parsed_parts, parts_data, overwrite
                            )
                        )

                        # データを保存
                        save_parts_data(result_parts)

                        # 結果を保存
                        st.session_state.csv_import_result = {
                            "success": success,
                            "skip": skip,
                            "error": error,
                            "duplicates": dup_list
                        }

                        st.rerun()

                else:
                    st.warning("⚠️ CSVファイルに有効な部品データが見つかりませんでした")

            except Exception as e:
                st.error(f"❌ CSVファイルの読み込み中にエラーが発生しました: {str(e)}")

        # インポート結果の表示
        if st.session_state.csv_import_result:
            result = st.session_state.csv_import_result
            st.markdown("---")
            st.markdown("#### 📊 インポート結果")

            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("✅ 成功", f"{result['success']} 件")
            with col2:
                st.metric("⏭️ スキップ", f"{result['skip']} 件")
            with col3:
                st.metric("❌ エラー", f"{result['error']} 件")

            if result["skip"] > 0:
                with st.expander("スキップした部品の詳細"):
                    for dup in result["duplicates"]:
                        st.markdown(f"- {dup['id']}: {dup['name']} (重複)")

            # ホームに戻るボタン
            if st.button("🏠 ホームに戻る", type="primary", key="home_after_import"):
                st.session_state.csv_import_result = None
                st.session_state.csv_parsed_parts = []
                st.query_params.clear()
                st.rerun()


def show_inspection_form_page(parts_data, preselected_part_id=None):
    """Display inspection form page"""
    st.title("📋 検査表入力")
    st.markdown("---")

    # 検査項目をテンプレートから読み込み
    inspection_items = load_inspection_template()

    # 基本情報入力
    info_col1, info_col2, info_col3 = st.columns(3)
    with info_col1:
        inspection_date = st.date_input("検査日", value=datetime.now())
    with info_col2:
        inspector_name = st.text_input("検査者名", placeholder="山田太郎")
    with info_col3:
        # 部品選択
        part_options = ["選択してください"] + [
            f"{p['id']} - {p['name']}" for p in parts_data
        ]

        # 事前に選択された部品がある場合、そのインデックスを見つける
        default_index = 0
        if preselected_part_id:
            for idx, option in enumerate(part_options):
                if option.startswith(f"{preselected_part_id} -"):
                    default_index = idx
                    break

        selected_part_for_inspection = st.selectbox(
            "対象部品", part_options, index=default_index
        )

    # 選択された部品の情報を取得
    selected_part_info = None
    if selected_part_for_inspection != "選択してください":
        part_id = selected_part_for_inspection.split(" - ")[0]
        selected_part_info = next(
            (p for p in parts_data if p["id"] == part_id), None
        )

    st.markdown("---")

    # 2カラムレイアウト：左に部品情報、右に検査入力
    left_col, right_col = st.columns([1, 2])

    # 左カラム：部品情報（固定表示）
    with left_col:
        st.markdown("### 📌 部品情報")

        if selected_part_info:
            st.markdown(f"**{selected_part_info['name']}**")
            st.caption(f"ID: {selected_part_info['id']}")

            # 検査箇所画像
            image_path = get_image_path(selected_part_info)
            if image_path:
                st.image(image_path, caption="検査箇所", width="stretch")
            else:
                st.markdown(
                    f"""
                    <div style="
                        background-color: #f0f0f0;
                        border: 1px dashed #ccc;
                        border-radius: 5px;
                        padding: 20px;
                        text-align: center;
                        color: #666;
                        font-size: 12px;
                    ">
                        🔍 {selected_part_info.get(
                            'image_description', '検査箇所'
                        )}
                    </div>
                    """,
                    unsafe_allow_html=True
                )

            # 検査項目
            st.markdown("#### ✅ 検査項目")
            for item in selected_part_info.get("inspection_items", []):
                st.markdown(f"- {item}")

            # 注意点
            st.markdown("#### ⚠️ 注意点")
            for caution in selected_part_info.get("cautions", []):
                st.warning(caution)

            # 保管場所
            st.markdown(
                f"**📍 保管場所:** {selected_part_info['storage']}"
            )
        else:
            st.info("👆 対象部品を選択すると、検査項目と注意点が表示されます")

    # 右カラム：検査入力フォーム
    with right_col:
        st.markdown("### 📝 測定値入力")
        st.caption(
            "💡 項目1,6は「OK」または「NG」を入力、"
            "項目2-5は数値を入力すると自動判定されます"
        )

        # 検査項目の入力フォーム
        results = []
        all_judgments = []

        for item in inspection_items:
            st.markdown(f"**{item['no']}. {item['item']}**")
            st.caption(f"判定基準: {item['criteria']}")

            col1, col2, col3 = st.columns([2, 1, 2])

            with col1:
                result = st.text_input(
                    "測定値/結果",
                    key=f"result_{item['no']}",
                    placeholder="OK/NG または 数値を入力"
                )

            # 自動判定
            auto_judgment = auto_judge(item["no"], result, item["criteria"])

            with col2:
                if auto_judgment:
                    # 自動判定結果を表示
                    if auto_judgment == "合格":
                        st.success(f"判定: {auto_judgment}")
                    else:
                        st.error(f"判定: {auto_judgment}")
                    judgment = auto_judgment
                else:
                    # 手動選択
                    judgment = st.selectbox(
                        "判定",
                        ["", "合格", "不合格"],
                        key=f"judgment_{item['no']}"
                    )

            with col3:
                note = st.text_input(
                    "備考",
                    key=f"note_{item['no']}",
                    placeholder="備考（任意）"
                )

            results.append({
                "no": item["no"],
                "item": item["item"],
                "criteria": item["criteria"],
                "result": result,
                "judgment": judgment,
                "note": note
            })

            if judgment:
                all_judgments.append(judgment)

            st.markdown("---")

    # 総合判定（自動計算）
    st.markdown("### 総合判定")

    # 全ての項目が判定済みかチェック
    all_items_judged = len(all_judgments) == len(inspection_items)

    if all_items_judged:
        # 一つでも不合格があれば総合不合格
        if "不合格" in all_judgments:
            overall_judgment = "不合格"
            st.error(
                f"🔴 総合判定: **{overall_judgment}**（不合格項目があります）"
            )
        else:
            overall_judgment = "合格"
            st.success(
                f"🟢 総合判定: **{overall_judgment}**（全項目合格）"
            )
    else:
        overall_judgment = ""
        st.warning("⚠️ 全ての検査項目を入力すると総合判定が表示されます")

    # PDF出力ボタン
    st.markdown("---")

    # 部品データを取得
    selected_part_data = None
    if selected_part_for_inspection != "選択してください":
        part_id = selected_part_for_inspection.split(" - ")[0]
        selected_part_data = next(
            (p for p in parts_data if p["id"] == part_id), None
        )

    # 全項目入力済み、かつ全て合格の場合のみボタンを有効化
    has_failure = "不合格" in all_judgments
    button_disabled = not (
        all_items_judged
        and overall_judgment == "合格"
        and inspector_name
        and selected_part_for_inspection != "選択してください"
    )

    if has_failure:
        st.error(
            "🚫 不合格項目があるためPDF出力できません。"
            "全ての項目を合格にしてください。"
        )
    elif button_disabled:
        st.info(
            "📝 全ての項目（検査日、検査者、対象部品、各検査項目）を"
            "入力し、全て合格するとPDF出力ボタンが有効になります。"
        )

    if st.button(
        "📄 PDFで出力",
        width="stretch",
        disabled=button_disabled
    ):
        # 検査データを作成
        inspection_data = {
            "date": inspection_date.strftime("%Y-%m-%d"),
            "inspector": inspector_name,
            "items": results,
            "overall_judgment": overall_judgment
        }

        # PDF生成
        pdf_bytes = generate_pdf(inspection_data, selected_part_data or {})

        # ダウンロードボタン
        st.download_button(
            label="📥 PDFをダウンロード",
            data=pdf_bytes,
            file_name=(
                f"inspection_{selected_part_data['id']}_"
                f"{inspection_date.strftime('%Y%m%d')}.pdf"
            ),
            mime="application/pdf",
            type="primary"
        )

        st.success(
            "✅ PDF出力の準備ができました！"
            "上のボタンからダウンロードしてください。"
        )


def show_edit_part_page(part_id, parts_data):
    """Display edit part page"""
    # Find the selected part
    part_data = next((p for p in parts_data if p["id"] == part_id), None)

    if not part_data:
        st.error(f"部品ID '{part_id}' が見つかりません。")
        if st.button("ホームに戻る"):
            st.query_params.clear()
            st.rerun()
        return

    st.title(f"✏️ 部品編集: {part_data['name']}")
    st.markdown("---")

    # 編集フォーム
    with st.form("edit_part_form"):
        col1, col2 = st.columns(2)

        with col1:
            edit_id = st.text_input(
                "部品ID *",
                value=part_data["id"],
                disabled=True,
                help="部品IDは変更できません"
            )
            edit_name = st.text_input(
                "部品名 *",
                value=part_data["name"],
                placeholder="例: 六角ボルト M12"
            )
            edit_category = st.text_input(
                "カテゴリ *",
                value=part_data["category"],
                placeholder="例: 締結部品"
            )
            edit_storage = st.text_input(
                "保管場所 *",
                value=part_data["storage"],
                placeholder="例: A棟-1F-棚番号A-15"
            )

        with col2:
            edit_inspection = st.text_area(
                "検査項目 *（1行に1項目）",
                value="\n".join(part_data.get("inspection_items", [])),
                placeholder="ねじ山の損傷確認\n頭部の変形確認\n表面の錆確認",
                height=100
            )
            edit_cautions = st.text_area(
                "注意点（1行に1項目）",
                value="\n".join(part_data.get("cautions", [])),
                placeholder="トルク管理が重要\n再使用回数に注意",
                height=100
            )
            edit_image_desc = st.text_input(
                "検査箇所イメージの説明",
                value=part_data.get("image_description", ""),
                placeholder="例: ボルト頭部・ねじ山部の検査ポイント"
            )

            # 必須製品の現在の値を整形
            current_products = []
            for product in part_data.get("required_products", []):
                product_line = f"{product['product_id']}|{product['product_name']}"
                if product.get('notes'):
                    product_line += f"|{product['notes']}"
                current_products.append(product_line)

            edit_required_products = st.text_area(
                "必須製品（任意、1行に1製品）",
                value="\n".join(current_products),
                placeholder="TUA60|TUA60 アセンブリ|主軸固定用\nTUA70|TUA70 ユニット|予備用",
                height=80,
                help="形式: 製品ID|製品名|用途（パイプ区切り）"
            )

        # 現在の画像を表示
        st.markdown("#### 現在の画像")
        image_path = get_image_path(part_data)
        if image_path:
            col_img1, col_img2 = st.columns([1, 2])
            with col_img1:
                st.image(image_path, caption="現在の画像", width=200)
            with col_img2:
                st.info("新しい画像をアップロードすると、現在の画像が置き換えられます。")
        else:
            st.info("現在、画像は登録されていません。")

        # 画像アップロード
        uploaded_image = st.file_uploader(
            "新しい検査箇所の画像（任意）",
            type=["png", "jpg", "jpeg"],
            help="PNG, JPG, JPEG形式の画像をアップロードできます"
        )

        # フォームボタン
        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            submitted = st.form_submit_button("💾 更新", width="stretch", type="primary")
        with col_btn2:
            cancel = st.form_submit_button("❌ キャンセル", width="stretch")

        if cancel:
            # 部品詳細ページに戻る
            st.query_params["view"] = "part_details"
            st.query_params["part_id"] = part_id
            st.rerun()

        if submitted:
            # バリデーション
            if not edit_name or not edit_category or not edit_storage:
                st.error("必須項目（*）を入力してください。")
            elif not edit_inspection.strip():
                st.error("検査項目を1つ以上入力してください。")
            else:
                # 必須製品のパース
                required_products = []
                if edit_required_products.strip():
                    for line in edit_required_products.split("\n"):
                        if line.strip():
                            parts_info = [p.strip() for p in line.split("|")]
                            if len(parts_info) >= 2:
                                product = {
                                    "product_id": parts_info[0],
                                    "product_name": parts_info[1],
                                    "notes": (
                                        parts_info[2]
                                        if len(parts_info) >= 3
                                        else ""
                                    )
                                }
                                required_products.append(product)

                # 更新データを作成
                updated_part = {
                    "id": edit_id,
                    "name": edit_name,
                    "category": edit_category,
                    "inspection_items": [
                        item.strip() for item in edit_inspection.split("\n")
                        if item.strip()
                    ],
                    "cautions": [
                        item.strip() for item in edit_cautions.split("\n")
                        if item.strip()
                    ] if edit_cautions.strip() else ["特になし"],
                    "storage": edit_storage,
                    "image_description": (
                        edit_image_desc if edit_image_desc else "検査箇所"
                    ),
                    "required_products": required_products
                }

                # JSONに保存（画像も含む）
                if update_part(part_id, updated_part, uploaded_image):
                    st.success(f"部品 '{edit_name}' を更新しました！")
                    # 部品詳細ページに自動的に戻る
                    st.query_params["view"] = "part_details"
                    st.query_params["part_id"] = part_id
                    st.rerun()
                else:
                    st.error("部品の更新に失敗しました。")


# ============================================================
# Main Routing Logic
# ============================================================

# Check which view to show based on query parameters
if current_view == "part_details" and selected_part_id_from_url:
    show_part_details_page(selected_part_id_from_url, parts_data)
elif current_view == "edit_part" and selected_part_id_from_url:
    show_edit_part_page(selected_part_id_from_url, parts_data)
elif current_view == "product_details" and selected_product_id_from_url:
    # Extract product name from the product ID
    product_name = None
    for part in parts_data:
        for product in part.get("required_products", []):
            if product["product_id"] == selected_product_id_from_url:
                product_name = product["product_name"]
                break
        if product_name:
            break

    if product_name:
        show_product_details_page(
            selected_product_id_from_url,
            product_name,
            parts_data
        )
    else:
        st.error(f"製品ID '{selected_product_id_from_url}' が見つかりません。")
        if st.button("ホームに戻る"):
            st.query_params.clear()
            st.rerun()
elif current_view == "add_part":
    show_add_part_page(parts_data)
elif current_view == "inspection_form":
    show_inspection_form_page(parts_data, preselected_part_id_for_inspection)
else:
    # Show main page
    # メインエリア
    st.title("🔍 部品検査箇所表示システム")
    st.markdown(
        "検査する部品を選択して、検査項目・注意点・保管場所を確認できます。"
    )

    # 製品フィルタが有効な場合は表示
    if selected_product != "すべて":
        st.info(
            f"🏭 **製品フィルタ適用中:** {selected_product} "
            f"に必要な部品のみを表示しています"
        )

    st.markdown("---")

    # 部品追加フォーム
    if st.session_state.show_add_form:
        st.subheader("➕ 新規部品登録")

        # タブで手動登録とCSV一括登録を切り替え
        tab1, tab2 = st.tabs(["✍️ 手動登録", "📁 CSV一括登録"])

        with tab1:
            # 手動登録フォーム
            with st.form("add_part_form"):
                col1, col2 = st.columns(2)

                with col1:
                    new_id = st.text_input("部品ID *", placeholder="例: BLT-002")
                    new_name = st.text_input("部品名 *", placeholder="例: 六角ボルト M12")
                    new_category = st.text_input("カテゴリ *", placeholder="例: 締結部品")
                    new_storage = st.text_input(
                        "保管場所 *", placeholder="例: A棟-1F-棚番号A-15"
                    )

                with col2:
                    new_inspection = st.text_area(
                        "検査項目 *（1行に1項目）",
                        placeholder="ねじ山の損傷確認\n頭部の変形確認\n表面の錆確認",
                        height=100
                    )
                    new_cautions = st.text_area(
                        "注意点（1行に1項目）",
                        placeholder="トルク管理が重要\n再使用回数に注意",
                        height=100
                    )
                    new_image_desc = st.text_input(
                        "検査箇所イメージの説明",
                        placeholder="例: ボルト頭部・ねじ山部の検査ポイント"
                    )
                    new_required_products = st.text_area(
                        "必須製品（任意、1行に1製品）",
                        placeholder="TUA60|TUA60 アセンブリ|主軸固定用\nTUA70|TUA70 ユニット|予備用",
                        height=80,
                        help="形式: 製品ID|製品名|用途（パイプ区切り）"
                    )

                # 画像アップロード
                uploaded_image = st.file_uploader(
                    "検査箇所の画像（任意）",
                    type=["png", "jpg", "jpeg"],
                    help="PNG, JPG, JPEG形式の画像をアップロードできます"
                )

                submitted = st.form_submit_button("登録", width="stretch")

                if submitted:
                    # バリデーション
                    if not new_id or not new_name or not new_category or not new_storage:
                        st.error("必須項目（*）を入力してください。")
                    elif any(part["id"] == new_id for part in parts_data):
                        st.error(f"部品ID '{new_id}' は既に存在します。")
                    elif not new_inspection.strip():
                        st.error("検査項目を1つ以上入力してください。")
                    else:
                        # 必須製品のパース
                        required_products = []
                        if new_required_products.strip():
                            for line in new_required_products.split("\n"):
                                if line.strip():
                                    parts_info = [p.strip() for p in line.split("|")]
                                    if len(parts_info) >= 2:
                                        product = {
                                            "product_id": parts_info[0],
                                            "product_name": parts_info[1],
                                            "notes": (
                                                parts_info[2]
                                                if len(parts_info) >= 3
                                                else ""
                                            )
                                        }
                                        required_products.append(product)

                        # 新規部品データを作成
                        new_part = {
                            "id": new_id,
                            "name": new_name,
                            "category": new_category,
                            "inspection_items": [
                                item.strip() for item in new_inspection.split("\n")
                                if item.strip()
                            ],
                            "cautions": [
                                item.strip() for item in new_cautions.split("\n")
                                if item.strip()
                            ] if new_cautions.strip() else ["特になし"],
                            "storage": new_storage,
                            "image_description": (
                                new_image_desc if new_image_desc else "検査箇所"
                            ),
                            "required_products": required_products
                        }

                        # JSONに保存（画像も含む）
                        save_part(new_part, uploaded_image)
                        st.success(f"部品 '{new_name}' を登録しました！")
                        st.session_state.show_add_form = False
                        st.rerun()

        with tab2:
            # CSV一括登録フォーム
            st.markdown("#### 📁 CSVファイルから部品を一括登録")
            st.caption(
                "CSVフォーマット: 2列目=品目、3列目=図番、4列目=品名。"
                "品目のみの行は製品カテゴリを表します。"
            )

            # セッション状態の初期化
            if "csv_parsed_parts" not in st.session_state:
                st.session_state.csv_parsed_parts = []
            if "csv_import_result" not in st.session_state:
                st.session_state.csv_import_result = None

            # CSVファイルアップロード
            uploaded_csv = st.file_uploader(
                "CSVファイルを選択",
                type=["csv"],
                help="部品情報が記載されたCSVファイルをアップロードしてください",
                key="csv_uploader"
            )

            if uploaded_csv is not None:
                try:
                    # CSVをパース
                    parsed_parts = parse_csv_file(uploaded_csv)
                    st.session_state.csv_parsed_parts = parsed_parts

                    if len(parsed_parts) > 0:
                        st.success(f"✅ {len(parsed_parts)} 件の部品データを読み込みました")

                        # プレビューテーブル
                        st.markdown("#### 📋 プレビュー")
                        preview_data = []
                        for part in parsed_parts[:10]:  # 最初の10件を表示
                            preview_data.append({
                                "部品ID": part["id"],
                                "部品名": part["name"],
                                "品目": part.get("item_type", ""),
                                "製品": (
                                    part["required_products"][0]["product_name"]
                                    if part["required_products"]
                                    else ""
                                )
                            })

                        st.dataframe(preview_data, use_container_width=True)

                        if len(parsed_parts) > 10:
                            st.caption(f"...他 {len(parsed_parts) - 10} 件")

                        # 重複チェック
                        unique_parts, duplicates = check_duplicates(
                            parsed_parts, parts_data
                        )

                        if duplicates:
                            st.warning(
                                f"⚠️ {len(duplicates)} 件の重複する部品IDがあります"
                            )
                            with st.expander("重複する部品ID一覧"):
                                for dup in duplicates:
                                    st.markdown(f"- {dup['id']}: {dup['name']}")

                        # インポート設定
                        st.markdown("#### ⚙️ インポート設定")
                        overwrite = st.checkbox(
                            "重複する部品を上書きする",
                            value=False,
                            help="チェックすると、既存の部品データを上書きします"
                        )

                        # インポートボタン
                        if st.button(
                            f"📥 {len(parsed_parts)} 件の部品をインポート",
                            type="primary",
                            width="stretch"
                        ):
                            # インポート実行
                            result_parts, success, skip, error, dup_list = (
                                import_parts_from_csv(
                                    parsed_parts, parts_data, overwrite
                                )
                            )

                            # データを保存
                            save_parts_data(result_parts)

                            # 結果を保存
                            st.session_state.csv_import_result = {
                                "success": success,
                                "skip": skip,
                                "error": error,
                                "duplicates": dup_list
                            }

                            st.rerun()

                    else:
                        st.warning("⚠️ CSVファイルに有効な部品データが見つかりませんでした")

                except Exception as e:
                    st.error(f"❌ CSVファイルの読み込み中にエラーが発生しました: {str(e)}")

            # インポート結果の表示
            if st.session_state.csv_import_result:
                result = st.session_state.csv_import_result
                st.markdown("---")
                st.markdown("#### 📊 インポート結果")

                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("✅ 成功", f"{result['success']} 件")
                with col2:
                    st.metric("⏭️ スキップ", f"{result['skip']} 件")
                with col3:
                    st.metric("❌ エラー", f"{result['error']} 件")

                if result["skip"] > 0:
                    with st.expander("スキップした部品の詳細"):
                        for dup in result["duplicates"]:
                            st.markdown(f"- {dup['id']}: {dup['name']} (重複)")

                # 結果をクリア
                if st.button("結果をクリアして新しいファイルをインポート"):
                    st.session_state.csv_import_result = None
                    st.session_state.csv_parsed_parts = []
                    st.rerun()

        st.markdown("---")

    # 部品カード一覧
    st.subheader("📋 部品一覧")


    # 部品カード一覧
    st.subheader("📋 部品一覧")

    if not filtered_parts:
        st.warning("該当する部品が見つかりません。検索条件を変更してください。")
    else:
        # 3列のグリッドレイアウト
        cols = st.columns(3)

        for idx, part in enumerate(filtered_parts):
            col_idx = idx % 3
            with cols[col_idx]:
                # カードのスタイル
                is_selected = st.session_state.selected_part == part["id"]
                border_color = "#1E88E5" if is_selected else "#ddd"
                bg_color = "#E3F2FD" if is_selected else "#fff"

                st.markdown(
                    f"""
                    <div style="
                        background-color: {bg_color};
                        border: 2px solid {border_color};
                        border-radius: 10px;
                        padding: 15px;
                        margin-bottom: 10px;
                        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                    ">
                        <div style="font-size: 12px; color: #666;">{part['id']}</div>
                        <div style="font-size: 18px; font-weight: bold; margin: 5px 0;">
                            {part['name']}
                        </div>
                        <div style="
                            display: inline-block;
                            background-color: #E8F5E9;
                            color: #2E7D32;
                            padding: 3px 10px;
                            border-radius: 15px;
                            font-size: 12px;
                        ">{part['category']}</div>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

                if st.button(
                    "詳細を見る",
                    key=f"btn_{part['id']}",
                    width="stretch"
                ):
                    # Keep current filters when navigating to details
                    st.query_params["view"] = "part_details"
                    st.query_params["part_id"] = part["id"]
                    if selected_product != "すべて":
                        st.query_params["selected_product"] = selected_product
                    if search_query:
                        st.query_params["search_query"] = search_query
                    if selected_category != "すべて":
                        st.query_params["selected_category"] = selected_category
                    st.rerun()


    # フッター
    st.markdown("---")
    st.markdown(
        """
        <div style="text-align: center; color: #666; font-size: 12px;">
            部品検査箇所表示システム v1.0
        </div>
        """,
        unsafe_allow_html=True
)
