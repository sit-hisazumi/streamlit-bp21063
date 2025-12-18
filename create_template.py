"""検査表テンプレートExcelファイルを作成するスクリプト"""
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ワークブック作成
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "検査表テンプレート"

# スタイル定義
header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=12, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)
center_align = Alignment(horizontal="center", vertical="center")

# タイトル行
ws.merge_cells("A1:F1")
ws["A1"] = "部品検査表"
ws["A1"].font = Font(bold=True, size=16)
ws["A1"].alignment = center_align

# 基本情報ヘッダー
ws["A3"] = "検査日"
ws["B3"] = ""
ws["C3"] = "検査者"
ws["D3"] = ""
ws["E3"] = "部品ID"
ws["F3"] = ""

for cell in ["A3", "C3", "E3"]:
    ws[cell].font = header_font
    ws[cell].fill = PatternFill(
        start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
    )

# 検査項目テーブルヘッダー（6項目のサンプル）
headers = ["No.", "検査項目", "判定基準", "測定値/結果", "判定", "備考"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=5, column=col, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# サンプル検査項目（6項目）
inspection_items = [
    {
        "no": 1,
        "item": "外観検査",
        "criteria": "傷・変形・錆なきこと",
        "result": "",
        "judgment": "",
        "note": ""
    },
    {
        "no": 2,
        "item": "寸法検査（長さ）",
        "criteria": "100±0.5mm",
        "result": "",
        "judgment": "",
        "note": ""
    },
    {
        "no": 3,
        "item": "寸法検査（幅）",
        "criteria": "50±0.3mm",
        "result": "",
        "judgment": "",
        "note": ""
    },
    {
        "no": 4,
        "item": "寸法検査（厚さ）",
        "criteria": "10±0.1mm",
        "result": "",
        "judgment": "",
        "note": ""
    },
    {
        "no": 5,
        "item": "硬度検査",
        "criteria": "HRC 58-62",
        "result": "",
        "judgment": "",
        "note": ""
    },
    {
        "no": 6,
        "item": "動作確認",
        "criteria": "スムーズに動作すること",
        "result": "",
        "judgment": "",
        "note": ""
    },
]

for row_idx, item in enumerate(inspection_items, 6):
    ws.cell(row=row_idx, column=1, value=item["no"]).border = thin_border
    ws.cell(row=row_idx, column=2, value=item["item"]).border = thin_border
    ws.cell(row=row_idx, column=3, value=item["criteria"]).border = thin_border
    ws.cell(row=row_idx, column=4, value=item["result"]).border = thin_border
    ws.cell(row=row_idx, column=5, value=item["judgment"]).border = thin_border
    ws.cell(row=row_idx, column=6, value=item["note"]).border = thin_border

    # 中央揃え（No.と判定）
    ws.cell(row=row_idx, column=1).alignment = center_align
    ws.cell(row=row_idx, column=5).alignment = center_align

# 総合判定
ws.merge_cells("A13:B13")
ws["A13"] = "総合判定"
ws["A13"].font = header_font
ws["A13"].fill = PatternFill(
    start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
)
ws["A13"].alignment = center_align
ws["A13"].border = thin_border
ws["B13"].border = thin_border

ws.merge_cells("C13:F13")
ws["C13"] = ""
ws["C13"].border = thin_border

# 列幅調整
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 25
ws.column_dimensions["D"].width = 15
ws.column_dimensions["E"].width = 10
ws.column_dimensions["F"].width = 20

# 保存
wb.save("templates/inspection_template.xlsx")
print("テンプレートを作成しました: templates/inspection_template.xlsx")
